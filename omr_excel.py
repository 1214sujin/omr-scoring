import pandas as pd
from openpyxl import load_workbook

def load_students_list(path):

    df = pd.read_excel(path, dtype=str, keep_default_na=False)

    std_list = {}

    for _, row in df.iterrows():

        sid = str(row['학번']).strip()

        name = str(row['성명']).strip()

        std_list[sid] = name

    return std_list


def load_correct_answer(path):

    df = pd.read_excel(path, dtype=str, keep_default_na=False)

    answer_key = {}
    score_table = {}

    for _, row in df.iterrows():

        if not row['정답'].strip():
            break

        q = int(row['문항']) - 1

        answers = [int(x) for x in str(row['정답']).replace(' ', '').split(',')]

        answer_key[q] = answers

        score_table[q] = float(row['배점'])

    return answer_key, score_table


def save_result(results, extra, std_list, answer_key, score_table, subject_name):

    INFO_OFFSET = 7
    TOTAL_ANSWERS = len(next(iter(results.values()))) # OMR 양식 상 문제 개수
    TOTAL_QUESTIONS = len(answer_key)

    data = []

    all_sids = set(std_list.keys()) | set(results.keys())

    for sid in all_sids:
        sname = std_list[sid] if sid in std_list else '*잘못된 학번'
        row = {'학번': sid, '이름': sname,
               '객관식': 0, '주관식': 0, '총점': 0, '석차': 0}
        score = 0

        # 채점
        if sid in results:
            for q, k in answer_key.items():
                a = results[sid][q]

                row[f'A{q+1}'] = ','.join(map(str, a))

                if set(a) == set(k):
                    row[f'{q+1}'] = 'o'
                    score += score_table[q]
                else:
                    row[f'{q+1}'] = 'x'

            row['객관식'] = score
            row['주관식'] = extra
            row['총점'] = score + extra
        else:
            for q in answer_key.keys():
                row[f'{q+1}'] = 'x'

        data.append(row)

    df = pd.DataFrame(data)
    
    df['석차'] = df['총점'].rank(method='min', ascending=False).astype(int)
    df = df.sort_values(by='석차').reset_index(drop=True)
    
    
    # 템플릿에 작성
    wb = load_workbook('omr_output_template.xlsx')
    ws = wb['Sheet1']

    start_row = 3  # 데이터 시작 행

    for i, row in df.iterrows():
        excel_row = start_row + i

        # 기본 정보
        ws.cell(row=excel_row, column=1, value=row['학번'])
        ws.cell(row=excel_row, column=2, value=row['이름'])
        ws.cell(row=excel_row, column=3, value=row['객관식'])
        ws.cell(row=excel_row, column=4, value=row['주관식'])
        ws.cell(row=excel_row, column=5, value=row['총점'])
        ws.cell(row=excel_row, column=6, value=row['석차'])

        # 정오표
        for q in range(TOTAL_QUESTIONS):
            ws.cell(row=excel_row, column=INFO_OFFSET+q, value=row[f'{q+1}'])

        # 답안
        offset = INFO_OFFSET + TOTAL_ANSWERS
        for q in range(TOTAL_QUESTIONS):
            ws.cell(row=excel_row, column=offset+q, value=row[f'A{q+1}'])

    out = f'결과출력/{subject_name}/{subject_name}_채점결과.xlsx'
    wb.save(out)
    
    print(f'결과 파일: {out}')
